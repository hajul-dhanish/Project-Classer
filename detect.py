import shutil
import cv2
import dlib
import os
import sys
import sqlite3


# ///////////////////////////////////////
# Global Variable
# //////////////////////////////////////

personGroupId = "test1"

key = "hidden"

BASE_URL = "https://hidden.api.cognitive.microsoft.com/face/v1.0"


# ///////////////////////////////////////
# Detection
# //////////////////////////////////////


# cam = cv2.VideoCapture(0)
detector = dlib.get_frontal_face_detector()
# img = cv2.imread(str(cam))
videoCaptureObject = cv2.VideoCapture(0)
result = True
while result:
    ret, frame = videoCaptureObject.read()
    cv2.imwrite("NewPicture.jpg", frame)

    hi = "NewPicture.jpg"
    # hi = "/home/dhanish/Music/SecondReview/upload/20211222_152506.jpg"
    result = False
videoCaptureObject.release()
cv2.destroyAllWindows()


img = cv2.imread(str(hi))

dets = detector(img, 1)
if not os.path.exists("./Cropped_faces"):
    os.makedirs("./Cropped_faces")
print("detected = {}".format(len(dets)))
for i, d in enumerate(dets):
    cv2.imwrite(
        "./Cropped_faces/face" + str(i + 1) + ".jpg",
        img[d.top() : d.bottom(), d.left() : d.right()],
    )

# ///////////////////////////////////////
# creating Excel sheet
# //////////////////////////////////////

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter, column_index_from_string
import time
import os
import sqlite3

# database connection
conn = sqlite3.connect("Face-DataBase")
c = conn.cursor()

# get current date
currentDate = time.strftime("%d_%m_%y")

# create a workbook and add a worksheet
if os.path.exists("./reports.xlsx"):
    wb = load_workbook(filename="reports.xlsx")
    sheet = wb["ECE-A"]
    # sheet[ord() + "1"]
    for col_index in range(1, 100):
        col = get_column_letter(col_index)
        if sheet.cell(row=1, column=1).value is None:
            col2 = get_column_letter(col_index - 1)
            # print sheet.cell('%s%s'% (col2, 1)).value
            if sheet.cell("%s%s" % (col2, 1)).value != currentDate:
                sheet["%s%s" % (col, 1)] = currentDate
            break
    # saving the file
    wb.save(filename="reports.xlsx")

else:
    wb = Workbook()
    dest_filename = "reports.xlsx"
    c.execute("SELECT * FROM Students ORDER BY Roll ASC")

    # creating worksheet and giving names to column
    ws1 = wb.active
    ws1.title = "ECE-A"
    ws1.append(("Roll Number", "Name", currentDate))
    ws1.append(("", "", ""))

    # entering students information from database
    while True:
        a = c.fetchone()
        if a == None:
            break
        else:
            ws1.append((a[2], a[1]))

    # saving the file
    wb.save(filename=dest_filename)


# ///////////////////////////////////////
# Identify
# //////////////////////////////////////


import cognitive_face as CF
import os, urllib
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell import Cell
import time
import requests
from requests.packages.urllib3.exceptions import InsecureRequestWarning

requests.packages.urllib3.disable_warnings(InsecureRequestWarning)


# get current date
currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename="reports.xlsx")
sheet = wb["ECE-A"]


def getDateColumn():
    for i in range(1, len(list(sheet.rows)[0]) + 1):
        col = get_column_letter(i)
        if sheet["%s%s" % (col, "1")].value == currentDate:
            return col


CF.Key.set(key)

CF.BaseUrl.set(BASE_URL)

connect = sqlite3.connect("Face-DataBase")
# c = connect.cursor()

attend = [0 for i in range(60)]

currentDir = os.path.dirname(os.path.abspath(__file__))
directory = os.path.join(currentDir, "Cropped_faces")
for filename in os.listdir(directory):
    if filename.endswith(".jpg"):
        imgurl = urllib.request.pathname2url(os.path.join(directory, filename))
        # imgurl = imgurl[3:]
        print("imgurl = {}".format(imgurl))
        res = CF.face.detect(imgurl)
        print("Res = {}".format(res))

        if len(res) < 1:
            print("No face detected.")
            continue

        faceIds = []
        for face in res:
            faceIds.append(face["faceId"])
        res = CF.face.identify(faceIds, personGroupId)
        print(filename)
        print("res = {}".format(res))

        for face in res:
            if not face["candidates"]:
                print("Unknown")
            else:
                personId = face["candidates"][0]["personId"]
                print("personid = {}".format(personId))
                # cmd =  + personId
                cur = connect.execute(
                    "SELECT * FROM Students WHERE personID = (?)", (personId,)
                )
                # print("cur = {}".format(cur))
                for row in cur:
                    print("Processing.......")
                    print("row = {}".format(row))
                    attend[int(row[0])] += 1
                # print("---------- " + row[1] + " recognized ----------")
        time.sleep(6)

for row in range(2, len(list(sheet.columns)[0]) + 1):
    rn = sheet.cell(row=row, column=1).value
    if rn is not None:
        print("rn = {}".format(rn))
        # rn = rn[-2:]
        # if attend[int(rn)] != 0:
        col = getDateColumn()
        print("col = {}".format(col))
        sheet["%s%s" % (col, str(row))] = 1

wb.save(filename="reports.xlsx")

shutil.rmtree("/Cropped_faces")
# os.remove("/home/dhanish/Music/SecondReview/Cropped_faces")

# currentDir = os.path.dirname(os.path.abspath(__file__))
# imgurl = urllib.pathname2url(os.path.join(currentDir, "1.jpg"))
# res = CF.face.detect(imgurl)
# faceIds = []
# for face in res:
#     faceIds.append(face["faceId"])

# res = CF.face.identify(faceIds, personGroupId)
# for face in res:
#     personName = CF.person.get(personGroupId, face["candidates"]["personId"])
#     print(personName)
# print(res)
